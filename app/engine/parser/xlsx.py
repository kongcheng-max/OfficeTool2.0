"""XLSX 解析器 — 基于 openpyxl"""

from typing import List

from openpyxl import load_workbook

from engine.parser.base import BaseParser, Chunk


class XLSXParser(BaseParser):
    name = "xlsx"
    supported_extensions = [".xlsx", ".xlsm"]
    supported_mime_types = [
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    ]

    async def parse(self, file_path: str, original_filename: str) -> List[Chunk]:
        chunks: List[Chunk] = []
        wb = load_workbook(file_path, read_only=True, data_only=True)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            # 将 Sheet 内容转为可读文本
            text_parts = [f"[Sheet: {sheet_name}]"]

            # 表头
            if rows:
                header = [str(c) if c is not None else "" for c in rows[0]]
                text_parts.append("表头: " + " | ".join(header))

            # 数据行（限制最多 500 行以避免单 Chunk 过大）
            for i, row in enumerate(rows[1:501], 1):
                values = [str(c) if c is not None else "" for c in row]
                if any(v.strip() for v in values):
                    text_parts.append(f"行{i}: " + " | ".join(values))

            if len(rows) > 501:
                text_parts.append(f"... (共 {len(rows)} 行，已截断前 500 行)")

            content = "\n".join(text_parts)

            chunks.append(Chunk(
                content=content,
                metadata={
                    "sheet": sheet_name,
                    "source": original_filename,
                    "parser_name": self.name,
                    "chunk_index": len(chunks),
                    "row_count": len(rows),
                },
                chunk_type="table",
            ))

        wb.close()
        return chunks
